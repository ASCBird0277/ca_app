from pathlib import Path
import json

try:
    import openpyxl  # type: ignore
except Exception as e:  # pragma: no cover
    print("OPENPYXL_IMPORT_ERROR", str(e))
    raise

FILES = [
    Path('data/Properties_geocoded.xlsx'),
    Path('data/Employee.xlsx'),
    Path('data/Positions.xlsx'),
]

for p in FILES:
    if not p.exists():
        print(f"FILE: {p} -> MISSING")
        continue
    try:
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = wb.active
        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        rows = []
        for _, row in zip(range(5), ws.iter_rows(min_row=2)):
            rows.append([c.value for c in row])
        print(f"FILE: {p}")
        print("HEADERS:", json.dumps(header, ensure_ascii=False))
        if rows:
            print("SAMPLE_ROW_1:", json.dumps(rows[0], ensure_ascii=False))
    except Exception as e:
        print(f"FILE: {p} ERROR: {e}")

